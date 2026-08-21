import json
import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, NamedTuple

from pydantic import Field, field_validator

from ..models import StrictModel
from ..source_types import SourceCommand, SourceContext, SourceDataError

CellValue = bool | int | float | str


def require_workbook_extension(value: str, command_name: str) -> str:
    """Requires the path to target an XLSX/XLSM workbook.

    Args:
        value: Workspace-relative path from verifier.json.
        command_name: Public source command name for the validation error.

    Returns:
        The unchanged path when it has a supported workbook extension.

    Raises:
        ValueError: If the path does not end in ``.xlsx`` or ``.xlsm``.
    """
    if Path(value).suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError(f"{command_name} requires a .xlsx or .xlsm file")
    return value


class XlsxExtractTextInput(StrictModel):
    """Input for xlsx.extract_text.

    Attributes:
        path: Workspace-relative XLSX or XLSM path.
    """

    path: str

    @field_validator("path")
    @classmethod
    def require_workbook_extension(cls, value: str) -> str:
        """Requires the path to target an XLSX/XLSM workbook."""
        return require_workbook_extension(value, "xlsx.extract_text")


class XlsxExtractTextOutput(StrictModel):
    """Output from xlsx.extract_text.

    Attributes:
        text: Extracted sheet names and non-empty cell rows.
    """

    text: str


class ExtractText(SourceCommand[XlsxExtractTextInput, XlsxExtractTextOutput]):
    """Extracts readable text from workbook sheets."""

    name = "extract_text"
    input_model = XlsxExtractTextInput
    output_model = XlsxExtractTextOutput

    def run(self, source_input: XlsxExtractTextInput, context: SourceContext) -> XlsxExtractTextOutput:
        """Runs XLSX text extraction.

        Args:
            source_input: Validated workbook extraction input.
            context: Source runtime context.

        Returns:
            Extracted workbook text.

        """
        resolved = context.resolve_path(source_input.path)
        return XlsxExtractTextOutput(text=extract_xlsx(resolved, context.max_content_chars))


class XlsxExtractFormulasInput(StrictModel):
    """Input for xlsx.extract_formulas.

    Attributes:
        path: Workspace-relative XLSX or XLSM path.
    """

    path: str

    @field_validator("path")
    @classmethod
    def require_workbook_extension(cls, value: str) -> str:
        """Requires the path to target an XLSX/XLSM workbook."""
        return require_workbook_extension(value, "xlsx.extract_formulas")


class XlsxExtractFormulasOutput(StrictModel):
    """Output from xlsx.extract_formulas.

    Attributes:
        text: One ``Sheet!Cell: =FORMULA`` line per formula cell in the
            workbook, or an empty string when the workbook holds none.
    """

    text: str


class ExtractFormulas(SourceCommand[XlsxExtractFormulasInput, XlsxExtractFormulasOutput]):
    """Lists every formula in a workbook, sheet-qualified, one per line."""

    name = "extract_formulas"
    input_model = XlsxExtractFormulasInput
    output_model = XlsxExtractFormulasOutput

    def run(self, source_input: XlsxExtractFormulasInput, context: SourceContext) -> XlsxExtractFormulasOutput:
        """Runs XLSX formula extraction.

        Args:
            source_input: Validated workbook extraction input.
            context: Source runtime context.

        Returns:
            Extracted workbook formula text.

        """
        resolved = context.resolve_path(source_input.path)
        return XlsxExtractFormulasOutput(text=extract_xlsx_formulas(resolved, context.max_content_chars))


class XlsxReadCellInput(StrictModel):
    """Input for xlsx.read_cell.

    Attributes:
        path: Workspace-relative XLSX or XLSM path.
        sheet: Workbook sheet name.
        cell: A1-style cell coordinate.
    """

    path: str
    sheet: str
    cell: str

    @field_validator("path")
    @classmethod
    def require_workbook_extension(cls, value: str) -> str:
        """Requires the path to target an XLSX/XLSM workbook."""
        return require_workbook_extension(value, "xlsx.read_cell")

    @field_validator("sheet")
    @classmethod
    def require_sheet_name(cls, value: str) -> str:
        """Requires a non-empty sheet name."""
        if not value.strip():
            raise ValueError("sheet must be non-empty")
        return value

    @field_validator("cell")
    @classmethod
    def require_cell_coordinate(cls, value: str) -> str:
        """Requires a valid A1-style cell coordinate."""
        from openpyxl.utils.cell import coordinate_to_tuple

        coordinate = value.strip().upper()
        if not coordinate:
            raise ValueError("cell must be an A1-style coordinate")
        try:
            row, column = coordinate_to_tuple(coordinate)
        except ValueError as exc:
            raise ValueError("cell must be an A1-style coordinate") from exc
        if row < 1 or column < 1:
            raise ValueError("cell must be an A1-style coordinate")
        return coordinate


class XlsxReadCellOutput(StrictModel):
    """Output from xlsx.read_cell.

    ``value`` stays declared first so a rubric judge's evidence block keeps the
    shape it had before ``formula`` and ``has_formula`` were added.

    Attributes:
        value: JSON-safe value read from the requested workbook cell. ``None``
            when the cell holds a formula whose result was never computed into
            the file — the normal state of a workbook written by openpyxl.
        formula: The cell's ``=``-prefixed formula. ``None`` for a cell holding
            no formula, and for the rare formula object that exposes no text.
        has_formula: Whether the cell holds a formula at all. Weaker than
            ``formula``: ``=1`` satisfies it, so prefer asserting on the
            formula text itself.
        number_format: The cell's stored Excel number format, ``"General"``
            when none was applied. This is the only evidence for a prompt that
            requires presentation — currency, a percentage, two decimals —
            because ``value`` reads 1234.5 whether the cell shows ``$1,234.50``
            or ``1234.5``. Grade it with ``contains`` or ``regex_match``.
    """

    value: CellValue | None
    formula: str | None
    has_formula: bool
    number_format: str


class ReadCell(SourceCommand[XlsxReadCellInput, XlsxReadCellOutput]):
    """Reads one populated cell from a workbook sheet."""

    name = "read_cell"
    input_model = XlsxReadCellInput
    output_model = XlsxReadCellOutput

    def run(self, source_input: XlsxReadCellInput, context: SourceContext) -> XlsxReadCellOutput:
        """Runs XLSX cell reading.

        Args:
            source_input: Validated workbook cell input.
            context: Source runtime context.

        Returns:
            The cell's JSON-safe value alongside its formula, if any.

        Raises:
            SourceDataError: If the sheet is missing, or a cell holding no
                formula is blank.
        """
        resolved = context.resolve_path(source_input.path)
        reading = read_xlsx_cell(resolved, source_input.sheet, source_input.cell)
        return XlsxReadCellOutput(
            value=reading.value,
            formula=reading.formula,
            has_formula=reading.has_formula,
            number_format=reading.number_format,
        )


class XlsxInspectModelInput(StrictModel):
    """Input for ``xlsx.inspect_model``.

    ``focus_terms`` must come from the task prompt. They make large workbooks
    inspectable without coupling a verifier to a golden workbook layout.
    """

    path: str
    focus_terms: list[str] = Field(default_factory=list, max_length=20)

    @field_validator("path")
    @classmethod
    def require_workbook_extension(cls, value: str) -> str:
        return require_workbook_extension(value, "xlsx.inspect_model")

    @field_validator("focus_terms")
    @classmethod
    def validate_focus_terms(cls, value: list[str]) -> list[str]:
        result: list[str] = []
        for term in value:
            cleaned = term.strip()
            if not cleaned:
                raise ValueError("focus_terms entries must be non-empty")
            if len(cleaned) > 160:
                raise ValueError("focus_terms entries must be at most 160 characters")
            if cleaned not in result:
                result.append(cleaned)
        return result


class ModelCell(StrictModel):
    """One non-empty cell exposed as structured workbook evidence."""

    coordinate: str
    value: CellValue | None
    formula: str | None
    has_formula: bool
    number_format: str
    locked: bool
    editable: bool
    references: list[str] = Field(default_factory=list)
    focus_terms: list[str] = Field(default_factory=list)


class FormulaDependency(StrictModel):
    """A formula cell's direct textual reference.

    ``resolved`` means the reference is an ordinary A1/range reference the
    extractor recognised. The raw reference remains available in every case so
    the rubric can avoid pretending unsupported Excel syntax was understood.
    """

    formula_cell: str
    reference: str
    resolved: bool


class ModelSheet(StrictModel):
    """Workbook sheet metadata and selected non-empty cells."""

    name: str
    hidden: bool
    protected: bool
    non_empty_cell_count: int
    formula_count: int
    cells: list[ModelCell] = Field(default_factory=list)


class XlsxInspectModelOutput(StrictModel):
    """Structured, bounded evidence about a workbook model.

    ``unresolved_references`` and ``broken_references`` answer different
    questions and must not be conflated. The first reports operand syntax this
    extractor does not model -- a named range, a structured table reference, a
    link to another workbook -- none of which is a defect. The second reports
    what Excel itself has lost: a formula still carrying ``#REF!`` because the
    range it pointed at was deleted. A workbook can have either without the
    other.

    Attributes:
        sheets: Per-sheet metadata and selected cells, bounded by the cap.
        formula_count: Formula cells across the workbook, never truncated.
        dependencies: Direct operands of the formulas that survived the cap.
        unresolved_references: Operands of the surviving formulas that are not
            ordinary A1/range references. Bounded by the cap, like the
            dependencies it is derived from.
        broken_references: ``Sheet!Cell`` ids of every formula carrying
            ``#REF!``, collected across the whole workbook rather than the
            surviving selection: a lost reference must not become invisible
            because the workbook was large.
        focus_terms: Prompt terms this run was bounded by.
        truncated: Whether any cell could not be represented within the cap.
    """

    sheets: list[ModelSheet]
    formula_count: int
    dependencies: list[FormulaDependency]
    unresolved_references: list[str]
    broken_references: list[str]
    focus_terms: list[str]
    truncated: bool


class InspectModel(SourceCommand[XlsxInspectModelInput, XlsxInspectModelOutput]):
    """Expose workbook structure, formulas, and effective input editability."""

    name = "inspect_model"
    input_model = XlsxInspectModelInput
    output_model = XlsxInspectModelOutput

    def run(
        self, source_input: XlsxInspectModelInput, context: SourceContext
    ) -> XlsxInspectModelOutput:
        resolved = context.resolve_path(source_input.path)
        return inspect_xlsx_model(
            resolved,
            context.max_content_chars,
            source_input.focus_terms,
        )


class XlsxInspectStructureInput(StrictModel):
    """Input for ``xlsx.inspect_structure``.

    Attributes:
        path: Workspace-relative XLSX or XLSM path.
    """

    path: str

    @field_validator("path")
    @classmethod
    def require_workbook_extension(cls, value: str) -> str:
        """Requires the path to target an XLSX/XLSM workbook."""
        return require_workbook_extension(value, "xlsx.inspect_structure")


class StructureSheet(StrictModel):
    """One worksheet's name, visibility, extent, and embedded object counts."""

    name: str
    hidden: bool
    max_row: int
    max_column: int
    chart_count: int
    image_count: int


class XlsxInspectStructureOutput(StrictModel):
    """Workbook structure expressed as counts and names.

    Deliberately distinct from ``xlsx.inspect_model``: that command reports
    formulas, dependencies and effective editability, is bounded by prompt
    ``focus_terms``, and is gated to rubric assertions because a workbook's
    model design is not a number. This one reports counts and names only, so
    "the workbook has four tabs" is a deterministic assertion instead of a
    judge call or one ``contains "[Sheet: <name>]"`` check per tab, which
    cannot express a count at all and needs the prompt to name every tab.

    Attributes:
        sheet_count: Worksheets in the workbook, hidden ones included.
        sheet_names: Worksheet names in workbook order.
        visible_sheet_count: Worksheets a reader sees without unhiding one.
        hidden_sheet_names: Names of the hidden and very-hidden worksheets.
        chart_count: Charts across the whole workbook.
        image_count: Embedded images across the whole workbook.
        sheets: Per-worksheet detail, in workbook order.
    """

    sheet_count: int
    sheet_names: list[str]
    visible_sheet_count: int
    hidden_sheet_names: list[str]
    chart_count: int
    image_count: int
    sheets: list[StructureSheet]


class InspectStructure(
    SourceCommand[XlsxInspectStructureInput, XlsxInspectStructureOutput]
):
    """Report worksheet tabs, extents, charts, and images as counts and names."""

    name = "inspect_structure"
    input_model = XlsxInspectStructureInput
    output_model = XlsxInspectStructureOutput

    def run(
        self, source_input: XlsxInspectStructureInput, context: SourceContext
    ) -> XlsxInspectStructureOutput:
        """Runs workbook structure inspection.

        Args:
            source_input: Validated workbook structure input.
            context: Source runtime context.

        Returns:
            Workbook structure as counts and names.
        """
        resolved = context.resolve_path(source_input.path)
        return inspect_xlsx_structure(resolved)


def inspect_xlsx_structure(path: Path) -> XlsxInspectStructureOutput:
    """Reads worksheet names, visibility, extents, and embedded object counts.

    Takes no content limit: every field is a count or a worksheet name, so this
    command returns no extracted document text to bound.

    Args:
        path: Resolved workbook path.

    Returns:
        Workbook structure as counts and names.
    """
    from openpyxl import load_workbook

    # NOT read_only, for the reason documented on extract_xlsx: read_only mode
    # trusts the stored <dimension> tag, which would make max_row/max_column
    # report the workbook's claim rather than its cells.
    wb = load_workbook(path, data_only=True)
    try:
        sheets = [
            StructureSheet(
                name=sheet_name,
                hidden=wb[sheet_name].sheet_state != "visible",
                max_row=int(wb[sheet_name].max_row or 0),
                max_column=int(wb[sheet_name].max_column or 0),
                chart_count=len(drawing_collection(wb[sheet_name], "_charts")),
                image_count=len(drawing_collection(wb[sheet_name], "_images")),
            )
            for sheet_name in wb.sheetnames
        ]
        return XlsxInspectStructureOutput(
            sheet_count=len(sheets),
            sheet_names=[sheet.name for sheet in sheets],
            visible_sheet_count=sum(1 for sheet in sheets if not sheet.hidden),
            hidden_sheet_names=[sheet.name for sheet in sheets if sheet.hidden],
            chart_count=sum(sheet.chart_count for sheet in sheets),
            image_count=sum(sheet.image_count for sheet in sheets),
            sheets=sheets,
        )
    finally:
        wb.close()


def drawing_collection(worksheet: Any, attribute: str) -> list[Any]:
    """Reads a worksheet's chart or image collection defensively.

    openpyxl exposes neither collection through a public accessor, so this
    reads the private attribute through ``getattr`` with a default. A future
    openpyxl that renames or drops it then degrades to a count of zero instead
    of raising ``AttributeError``, which is absent from ``agent_failure_types``
    and would turn one structural check into a whole-run infrastructure error.

    Args:
        worksheet: openpyxl worksheet.
        attribute: Private collection attribute name.

    Returns:
        The collection's members, or an empty list when it is absent.
    """
    collection = getattr(worksheet, attribute, None)
    return list(collection) if collection else []


def extract_xlsx(path: Path, limit: int) -> str:
    """Extracts sheet names and non-empty cell rows from a workbook.

    Args:
        path: Resolved workbook path.
        limit: Maximum number of characters to return.

    Returns:
        Extracted workbook text capped to ``limit`` characters.
    """
    from openpyxl import load_workbook

    # NOT read_only: read_only mode trusts the stored <dimension> tag, which can
    # be wrong/narrow and silently truncate rows. Normal mode scans real cells.
    wb = load_workbook(path, data_only=True)
    parts: list[str] = []
    total = 0
    try:
        for sheet_name in wb.sheetnames:
            total = append_limited(parts, f"[Sheet: {sheet_name}]", total, limit)
            if total >= limit:
                return "\n".join(parts)[:limit]
            for row in wb[sheet_name].iter_rows(values_only=True):
                cells = [str(cell) if cell is not None else "" for cell in row]
                if any(cell.strip() for cell in cells):
                    total = append_limited(parts, " | ".join(cells), total, limit)
                    if total >= limit:
                        return "\n".join(parts)[:limit]
        return "\n".join(parts)[:limit]
    finally:
        wb.close()


def inspect_xlsx_model(
    path: Path, limit: int, focus_terms: list[str]
) -> XlsxInspectModelOutput:
    """Extract bounded structural evidence without evaluating workbook formulas.

    Workbook formulas are data, not code: this routine never evaluates them
    and never opens a mutable copy of the artifact. It records direct formula
    references and effective Excel editability so an LLM rubric can assess
    prompt-grounded model design without depending on hard-coded cell names.
    """
    from openpyxl import load_workbook

    # Keep enough space for sheet summaries even when callers configure a
    # small source limit. The output still marks itself truncated whenever a
    # cell could not be represented.
    char_limit = max(limit, 1_000)
    wb_formula = load_workbook(path, data_only=False)
    wb_cached = load_workbook(path, data_only=True)
    lowered_terms = [(term, term.casefold()) for term in focus_terms]
    sheet_meta: dict[str, dict[str, Any]] = {}
    candidates: list[tuple[int, int, int, int, str, ModelCell]] = []
    dependencies_by_formula: dict[str, list[FormulaDependency]] = {}
    unresolved_by_formula: dict[str, list[str]] = {}
    broken_references: list[str] = []

    try:
        # Identify focus matches first so immediate label/value neighbourhoods
        # survive evidence capping even in large source-data tabs.
        focus_neighbours: set[tuple[str, int, int]] = set()
        for sheet_name in wb_formula.sheetnames:
            worksheet = wb_formula[sheet_name]
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is None and cell.data_type != "f":
                        continue
                    haystack = str(formula_text(cell.value) or cell.value or "").casefold()
                    if any(term_lower in haystack for _term, term_lower in lowered_terms):
                        for row_offset in range(-1, 2):
                            for column_offset in range(-1, 2):
                                neighbour_row = cell.row + row_offset
                                neighbour_column = cell.column + column_offset
                                if neighbour_row > 0 and neighbour_column > 0:
                                    focus_neighbours.add((sheet_name, neighbour_row, neighbour_column))

        for sheet_index, sheet_name in enumerate(wb_formula.sheetnames):
            worksheet = wb_formula[sheet_name]
            cached_sheet = wb_cached[sheet_name]
            protected = bool(worksheet.protection.sheet)
            non_empty_count = 0
            formula_count = 0
            sheet_meta[sheet_name] = {
                "index": sheet_index,
                "hidden": worksheet.sheet_state != "visible",
                "protected": protected,
                "non_empty_cell_count": 0,
                "formula_count": 0,
            }
            for row in worksheet.iter_rows():
                for cell in row:
                    is_formula = cell.data_type == "f"
                    if cell.value is None and not is_formula:
                        continue
                    non_empty_count += 1
                    formula = formula_text(cell.value) if is_formula else None
                    if is_formula:
                        formula_count += 1
                    cached_value = cached_sheet[cell.coordinate].value if is_formula else cell.value
                    safe_value = None if cached_value is None else json_safe_cell_value(cached_value)
                    matched_terms = [
                        term
                        for term, term_lower in lowered_terms
                        if term_lower in str(formula or safe_value or "").casefold()
                    ]
                    formula_id = f"{qualify_sheet_name(sheet_name)}!{cell.coordinate}"
                    # Collected here, in the whole-workbook scan, rather than
                    # from the surviving selection below: a deleted range is a
                    # defect whether or not its formula fits in the evidence.
                    if formula and "#REF!" in formula:
                        broken_references.append(formula_id)
                    refs = formula_references(formula) if formula else []
                    dependencies = [
                        FormulaDependency(
                            formula_cell=formula_id,
                            reference=reference,
                            resolved=is_resolved_formula_reference(reference),
                        )
                        for reference in refs
                    ]
                    if dependencies:
                        dependencies_by_formula[formula_id] = dependencies
                        unresolved_by_formula[formula_id] = [
                            item.reference for item in dependencies if not item.resolved
                        ]
                    locked = bool(cell.protection.locked)
                    model_cell = ModelCell(
                        coordinate=cell.coordinate,
                        value=safe_value,
                        formula=formula,
                        has_formula=is_formula,
                        number_format=str(cell.number_format or "General"),
                        locked=locked,
                        editable=not (protected and locked),
                        references=refs,
                        focus_terms=matched_terms,
                    )
                    # Prompt matches and their immediate label/value/formula
                    # neighbourhood must survive truncation ahead of unrelated
                    # formulas. Within the remaining evidence, formulas are
                    # stronger than ordinary content.
                    focused = bool(
                        matched_terms
                        or (sheet_name, cell.row, cell.column) in focus_neighbours
                    )
                    priority = 0 if focused else (1 if is_formula else 2)
                    candidates.append((priority, sheet_index, cell.row, cell.column, sheet_name, model_cell))
            sheet_meta[sheet_name]["non_empty_cell_count"] = non_empty_count
            sheet_meta[sheet_name]["formula_count"] = formula_count

        candidates.sort(key=lambda item: item[:4])
        selected_by_sheet: dict[str, list[ModelCell]] = {name: [] for name in wb_formula.sheetnames}
        selected_formula_ids: set[str] = set()
        total = 0
        truncated = False
        for _priority, _sheet_index, _row, _column, sheet_name, model_cell in candidates:
            estimated = len(json.dumps(model_cell.model_dump(), default=str, ensure_ascii=False)) + 1
            if total + estimated > char_limit:
                truncated = True
                continue
            selected_by_sheet[sheet_name].append(model_cell)
            total += estimated
            if model_cell.has_formula:
                selected_formula_ids.add(f"{qualify_sheet_name(sheet_name)}!{model_cell.coordinate}")

        dependencies = [
            dependency
            for formula_id in sorted(selected_formula_ids)
            for dependency in dependencies_by_formula.get(formula_id, [])
        ]
        unresolved = [
            reference
            for formula_id in sorted(selected_formula_ids)
            for reference in unresolved_by_formula.get(formula_id, [])
        ]
        sheets = [
            ModelSheet(
                name=sheet_name,
                hidden=sheet_meta[sheet_name]["hidden"],
                protected=sheet_meta[sheet_name]["protected"],
                non_empty_cell_count=sheet_meta[sheet_name]["non_empty_cell_count"],
                formula_count=sheet_meta[sheet_name]["formula_count"],
                cells=selected_by_sheet[sheet_name],
            )
            for sheet_name in wb_formula.sheetnames
        ]
        return XlsxInspectModelOutput(
            sheets=sheets,
            formula_count=sum(item["formula_count"] for item in sheet_meta.values()),
            dependencies=dependencies,
            unresolved_references=unresolved,
            broken_references=broken_references,
            focus_terms=focus_terms,
            truncated=truncated,
        )
    finally:
        wb_cached.close()
        wb_formula.close()


def formula_references(formula: str) -> list[str]:
    """Return direct range/name operands from an Excel formula conservatively."""
    if not formula:
        return []
    try:
        from openpyxl.formula import Tokenizer

        tokens = Tokenizer(formula).items
    except Exception:
        return []
    references: list[str] = []
    for token in tokens:
        if token.type == "OPERAND" and token.subtype == "RANGE":
            reference = str(token.value).strip()
            if reference and reference not in references:
                references.append(reference)
    return references


_A1_REFERENCE = re.compile(
    r"^(?:(?:'[^']*(?:''[^']*)*'|[A-Za-z_][A-Za-z0-9_. ]*)!)?\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?$"
)


def is_resolved_formula_reference(reference: str) -> bool:
    """Whether a formula operand is an ordinary internal A1/range reference."""
    return bool(_A1_REFERENCE.fullmatch(reference.strip()))


def extract_xlsx_formulas(path: Path, limit: int) -> str:
    """Extracts every formula in a workbook as sheet-qualified lines.

    Each line is ``Sheet!Cell: =FORMULA`` so it stands alone and an author can
    assert on it with ``contains`` without depending on the order sheets happen
    to be listed in. An array formula gains a trailing ``[array <ref>]``.

    A workbook with no formulas yields an empty string rather than an error, so
    a ``not_contains`` assertion can pass for it.

    Args:
        path: Resolved workbook path.
        limit: Maximum number of characters to return.

    Returns:
        Extracted formula text capped to ``limit`` characters.
    """
    from openpyxl import load_workbook

    # data_only=False is the whole point: the cached-value load returns the
    # formula's last computed result, which is what extract_xlsx already shows.
    # NOT read_only, for the same reason extract_xlsx is not: a wrong stored
    # <dimension> would silently truncate the scan.
    wb = load_workbook(path, data_only=False)
    parts: list[str] = []
    total = 0
    try:
        for sheet_name in wb.sheetnames:
            qualifier = qualify_sheet_name(sheet_name)
            for row in wb[sheet_name].iter_rows():
                for cell in row:
                    if cell.data_type != "f":
                        continue
                    formula = formula_text(cell.value)
                    if formula is None:
                        continue
                    line = f"{qualifier}!{cell.coordinate}: {formula}"
                    array_ref = getattr(cell.value, "ref", None)
                    if isinstance(array_ref, str):
                        line = f"{line} [array {array_ref}]"
                    total = append_limited(parts, line, total, limit)
                    if total >= limit:
                        return "\n".join(parts)[:limit]
        return "\n".join(parts)[:limit]
    finally:
        wb.close()


def formula_text(value: Any) -> str | None:
    """Returns a cell's formula as ``=``-prefixed text.

    Args:
        value: Raw value of a cell whose ``data_type`` is ``"f"``. Usually the
            formula string, but an array formula arrives as an ``ArrayFormula``
            object carrying its source in ``.text``.

    Returns:
        The formula text, or ``None`` for a formula object that exposes no
        text. ``DataTableFormula`` is the real case: it has no ``.text``, and
        ``AttributeError`` is not an agent-output failure, so reading the
        attribute unguarded would fail the whole run instead of one check.
    """
    text = value if isinstance(value, str) else getattr(value, "text", None)
    if not isinstance(text, str):
        return None
    return text if text.startswith("=") else f"={text}"


def qualify_sheet_name(sheet_name: str) -> str:
    """Quotes a sheet name for ``Sheet!Cell`` only when Excel's syntax requires it.

    ``openpyxl.utils.quote_sheetname`` quotes unconditionally, which would
    render the common case as ``'Summary'!B8`` and leave an author guessing at
    the fragment to assert on.

    Args:
        sheet_name: Workbook sheet name.

    Returns:
        The sheet name, quoted and escaped only when it holds a character that
        would otherwise be ambiguous in an A1-style reference.
    """
    if not any(character in sheet_name for character in " '!"):
        return sheet_name
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'"


class CellReading(NamedTuple):
    """One workbook cell as ``xlsx.read_cell`` reports it.

    Attributes:
        value: JSON-safe cell value, or ``None`` for a formula with no result
            computed into the file.
        formula: ``=``-prefixed formula text, or ``None``.
        has_formula: Whether the cell holds a formula.
        number_format: Stored Excel number format, ``"General"`` when unset.
    """

    value: CellValue | None
    formula: str | None
    has_formula: bool
    number_format: str


def read_xlsx_cell(path: Path, sheet: str, cell: str) -> CellReading:
    """Reads one cell from a workbook, with its formula when it has one.

    A formula cell never raises for blankness. A workbook written by openpyxl
    stores no computed result, so requiring one punished exactly the model that
    did the right thing: writing ``=SUM(A1:A2)`` failed while pasting ``30``
    passed. Its ``value`` reports ``None`` and the caller grades ``formula``.

    Args:
        path: Resolved workbook path.
        sheet: Sheet name to read from.
        cell: A1-style cell coordinate to read.

    Returns:
        The cell's value, formula, and whether it holds one.

    Raises:
        SourceDataError: If the sheet is missing, or a cell holding no formula
            is blank.
    """
    from openpyxl import load_workbook

    # NOT read_only: random cell access under a wrong stored <dimension> can read
    # blank for a populated cell. Normal mode reads the real value.
    wb = load_workbook(path, data_only=False)
    try:
        if sheet not in wb.sheetnames:
            raise SourceDataError(f"xlsx.read_cell sheet not found: {sheet}")
        target = wb[sheet][cell]
        number_format = str(target.number_format or "General")
        if target.data_type != "f":
            # A cell holding no formula reads identically under either load
            # mode, so the cached-value load below would find nothing new.
            # `is None`, never `not value`: 0, False and "" are real values.
            if target.value is None:
                raise SourceDataError(f"xlsx.read_cell cell is blank: {sheet}!{cell}")
            return CellReading(
                value=json_safe_cell_value(target.value),
                formula=None,
                has_formula=False,
                number_format=number_format,
            )
        formula = formula_text(target.value)
    finally:
        wb.close()

    return CellReading(
        value=read_cached_cell_value(path, sheet, cell),
        formula=formula,
        has_formula=True,
        number_format=number_format,
    )


def read_cached_cell_value(path: Path, sheet: str, cell: str) -> CellValue | None:
    """Reads the result a spreadsheet application last computed into a cell.

    Args:
        path: Resolved workbook path.
        sheet: Sheet name to read from.
        cell: A1-style cell coordinate to read.

    Returns:
        The cached JSON-safe value, or ``None`` when the file carries none.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    try:
        raw_value = wb[sheet][cell].value
        return None if raw_value is None else json_safe_cell_value(raw_value)
    finally:
        wb.close()


def json_safe_cell_value(value: Any) -> CellValue:
    """Converts an OpenPyXL cell value into a JSON-safe scalar."""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, str)):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (date, time)):
        return value.isoformat()
    return str(value)


def append_limited(parts: list[str], text: str, total: int, limit: int) -> int:
    """Appends text and returns the updated character count.

    Args:
        parts: Text fragments collected so far.
        text: Text fragment to append.
        total: Current approximate character count.
        limit: Maximum desired character count.

    Returns:
        Updated approximate character count.
    """
    parts.append(text)
    return total + len(text) + 1


COMMANDS = (
    ExtractText(),
    ExtractFormulas(),
    ReadCell(),
    InspectModel(),
    InspectStructure(),
)
