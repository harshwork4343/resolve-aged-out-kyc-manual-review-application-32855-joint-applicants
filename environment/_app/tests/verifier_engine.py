"""Shared OBI verifier engine — extracted from benchmark_runner.py.
Regenerate with: python scripts/sync_verifier_engine.py

LOCAL EDITS (re-apply after any regeneration):
  1. task_inputs / skill_tools imported plainly — both ship alongside this
     file, so a missing module fails loudly instead of silently falling back.
  2. _build_trace_evidence includes each call's result (clipped), not just
     {name, args}; paired with load_agent_trace's result join + ok in
     test_outputs.py. Without it, trace rubrics about a FINDING can't pass.
"""

from __future__ import annotations

import asyncio
import json
import logging
import os
import posixpath
import re
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import shutil
import tempfile
from pathlib import Path

import httpx

from task_inputs import INPUT_DIR_NAME
from skill_tools import list_workspace_files

logger = logging.getLogger(__name__)

def _env_int(name: str, default: int) -> int:
    raw = os.environ.get(name, "").strip()
    if not raw:
        return default
    try:
        return max(0, int(raw))
    except ValueError:
        return default
def _top_level_json_objects(s: str) -> List[Dict[str, Any]]:
    """Parseable top-level JSON objects embedded in ``s``, in position order.

    For every ``{`` we do a string- and escape-aware brace match to its close
    and try to parse that span. Objects fully contained within an already
    accepted span are skipped, so a nested sub-object is never returned in
    place of its enclosing root. Scanning from every ``{`` (rather than a
    single pass) keeps a stray brace or quote in surrounding prose from
    derailing detection of the real object.
    """
    objs: List[Dict[str, Any]] = []
    accepted: List[tuple] = []  # (start, end) of accepted top-level spans
    n = len(s)
    for start in range(n):
        if s[start] != "{":
            continue
        if any(a <= start < b for a, b in accepted):
            continue
        depth = 0
        in_str = False
        esc = False
        for j in range(start, n):
            ch = s[j]
            if in_str:
                if esc:
                    esc = False
                elif ch == "\\":
                    esc = True
                elif ch == '"':
                    in_str = False
                continue
            if ch == '"':
                in_str = True
            elif ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    try:
                        obj = json.loads(s[start : j + 1])
                        if isinstance(obj, dict):
                            objs.append(obj)
                            accepted.append((start, j + 1))
                    except json.JSONDecodeError:
                        pass
                    break
    return objs


def _walk_json_objects(node: Any):
    """Yield every dict nested anywhere inside a JSON value (self first)."""
    if isinstance(node, dict):
        yield node
        for value in node.values():
            yield from _walk_json_objects(value)
    elif isinstance(node, list):
        for value in node:
            yield from _walk_json_objects(value)


def _resolve_dotted_path(obj: Any, path: str) -> Tuple[bool, Any]:
    """Resolve "a.b.0.c" against nested dicts/lists. Returns (found, value)."""
    current = obj
    for segment in path.split("."):
        if isinstance(current, list):
            try:
                index = int(segment)
            except ValueError:
                return False, None
            if not -len(current) <= index < len(current):
                return False, None
            current = current[index]
        elif isinstance(current, dict):
            if segment not in current:
                return False, None
            current = current[segment]
        else:
            return False, None
    return True, current


def _extract_json_from_text(text: str) -> Optional[Dict[str, Any]]:
    """Parse the last JSON object from model text (fenced blocks or inline braces)."""
    if not text or not text.strip():
        return None
    stripped = text.strip()
    candidates: List[Dict[str, Any]] = []

    for block in re.findall(r"```(?:json)?\s*([\s\S]*?)```", stripped, re.I):
        try:
            obj = json.loads(block.strip())
            if isinstance(obj, dict):
                candidates.append(obj)
        except json.JSONDecodeError:
            continue

    try:
        obj = json.loads(stripped)
        if isinstance(obj, dict):
            candidates.append(obj)
    except json.JSONDecodeError:
        pass

    # Inline objects: string/escape-aware matching so a brace inside a string
    # value does not close early, and top-level only so a nested sub-object is
    # never returned in place of its root. Last by position wins.
    candidates.extend(_top_level_json_objects(stripped))

    if candidates:
        return candidates[-1]
    return None


def _prefer_json_answer_text(raw: str) -> str:
    """If the model wrapped JSON in prose/markdown, store the parsed object only."""
    parsed = _extract_json_from_text(raw)
    if parsed is not None:
        return json.dumps(parsed, ensure_ascii=False)
    return raw


def _parse_simple_json_path(path: str) -> List[str | int]:
    """Parse the small JSONPath subset used by legacy task configs.

    Supported forms are root-relative dotted paths with optional integer list
    indexes, e.g. ``$.warehouse.accounts_total`` or ``$.rows[0].count``.
    """
    if not isinstance(path, str) or not path.startswith("$"):
        raise ValueError(f"JSON path must start with '$': {path!r}")
    if path == "$":
        return []

    tokens: List[str | int] = []
    i = 1
    while i < len(path):
        if path[i] == ".":
            i += 1
            start = i
            while i < len(path) and path[i] not in ".[":
                i += 1
            key = path[start:i]
            if not key:
                raise ValueError(f"empty key in JSON path: {path!r}")
            tokens.append(key)
            continue
        if path[i] == "[":
            end = path.find("]", i)
            if end == -1:
                raise ValueError(f"unterminated index in JSON path: {path!r}")
            raw_index = path[i + 1 : end].strip()
            if not raw_index.isdigit():
                raise ValueError(f"only non-negative integer indexes are supported: {path!r}")
            tokens.append(int(raw_index))
            i = end + 1
            continue
        raise ValueError(f"unsupported JSON path syntax near {path[i:]!r}")
    return tokens


def _get_simple_json_path(data: Any, path: str) -> tuple[bool, Any]:
    """Return ``(found, value)`` for the runner's small JSONPath subset."""
    try:
        tokens = _parse_simple_json_path(path)
    except ValueError:
        raise

    current = data
    for token in tokens:
        if isinstance(token, int):
            if not isinstance(current, list) or token >= len(current):
                return False, None
            current = current[token]
            continue
        if not isinstance(current, dict) or token not in current:
            return False, None
        current = current[token]
    return True, current
def _normalize_ai_content(response: Any) -> str:
    """Plain text from AIMessage (string or provider content blocks)."""
    content = getattr(response, "content", None)
    if content is None:
        return ""
    if isinstance(content, list):
        parts: list[str] = []
        for block in content:
            if isinstance(block, dict):
                btype = str(block.get("type", ""))
                text = block.get("text") or block.get("content")
                if text is not None:
                    label = f"[{btype}] " if btype else ""
                    parts.append(f"{label}{text}")
                elif "reasoning" in btype.lower():
                    parts.append(_safe_json_preview(block, 4000))
            elif isinstance(block, str):
                parts.append(block)
        return "\n".join(parts).strip()
    return str(content).strip()


# ---------------------------------------------------------------------------
# Trace evidence for ``rubric_check`` (target="trace")
# ---------------------------------------------------------------------------
# Trace rubrics used to get ``{name, args}`` only — the judge saw the call but
# never its result, so any rubric about a *finding* was ungradeable. Results are
# included now, clipped per call so one schema dump can't eat the whole budget.
_TRACE_RESULT_MAX_CHARS = int(os.environ.get("RUBRIC_TRACE_RESULT_CHARS", "4000"))
_TRACE_RESULT_MIN_CHARS = int(os.environ.get("RUBRIC_TRACE_RESULT_MIN_CHARS", "200"))
_TRACE_EVIDENCE_MAX_CHARS = int(os.environ.get("RUBRIC_TRACE_EVIDENCE_CHARS", "200000"))


def _clip_json_value(value: Any, limit: int) -> str:
    """Compact-serialize *value* and clip it to *limit* chars, flagging the cut."""
    try:
        text = json.dumps(value, separators=(",", ":"), default=str)
    except (TypeError, ValueError):
        text = str(value)
    if len(text) <= limit:
        return text
    return text[:limit].rstrip() + "… [message truncated]"


def _extract_agent_messages(path: Path) -> str:
    """Every natural-language agent message, chronological and length-bounded.

    ``_extract_final_output`` returns only the LAST assistant message, which
    misses a value the model stated mid-conversation before a trailing
    pleasantry (e.g. a "Staff headcount: 13" report followed by "All done!").
    Rubric items whose intent is "at any point in the conversation" need the
    full spoken transcript, so this collects the text of every agent step.

    To keep the judge prompt bounded regardless of run length, each message is
    clipped to ``_AGENT_TRANSCRIPT_TURN_MAX_CHARS`` and the joined transcript is
    kept within ``_AGENT_TRANSCRIPT_MAX_CHARS`` tail-first (a final report or
    summary is almost always near the end). Tool observations are never mixed
    in. Returns "" for a flat / absent / unreadable trace, exactly like
    ``_extract_final_output``.
    """
    if not path.exists():
        return ""
    try:
        data = load_json(path)
    except (ValueError, OSError):
        return ""
    if not isinstance(data, dict):
        return ""
    steps = data.get("steps")
    if not isinstance(steps, list):
        return ""
    messages: list[str] = []
    for step in steps:
        if not isinstance(step, dict):
            continue
        if (step.get("source") or step.get("role")) in ("agent", "assistant"):
            text = _finish_action_message(step) or _step_message_text(step.get("message"))
            if text and text.strip():
                messages.append(_clip_turn(text, _AGENT_TRANSCRIPT_TURN_MAX_CHARS))
    if not messages:
        return ""
    # Tail-biased budget: keep the most recent messages within the char cap,
    # but always keep at least the final message.
    kept: list[str] = []
    total = 0
    dropped = False
    for m in reversed(messages):
        add = len(m) + 2  # account for the "\n\n" separator
        if kept and total + add > _AGENT_TRANSCRIPT_MAX_CHARS:
            dropped = True
            break
        kept.append(m)
        total += add
    kept.reverse()
    transcript = "\n\n".join(kept)
    if dropped:
        transcript = (
            "[… earlier assistant messages omitted to bound length …]\n\n" + transcript
        )
    return transcript


# ---------------------------------------------------------------------------
# benchmark_runner verifier engine — VENDORED VERBATIM + sandbox adapters
# ---------------------------------------------------------------------------
#
# The ``VerifierConfig`` dataclass + ``VerifierEngine`` class below are a
# VERBATIM copy of ``benchmark_runner.py``'s verifier layer (its
# ``execute_verifier`` / ``_execute_*_verifier`` / ``_judge_rubric`` /
# ``_compare_with_llm`` / ``_extract_value_from_sql_result`` / ``_compare_values``
# / ``_diff_json_expected`` methods) — NOT a hand re-port. When the batch enables
# ``agentic_llm_as_judge`` the engine OWNS THE WHOLE SCORE: it runs its typed
# verifier list and writes benchmark_runner's native result (``verification_results``
# + ``verification_summary`` + ``statistics``).
#
# The verbatim async, LangChain-and-live-gym code runs unchanged via three thin
# sandbox adapters:
#   * ``_ProxyProtocolClient.get_state(verify_queries=[q])`` wraps the sync
#     ``query_state(gym, q)`` proxy call into the ``{success, data:
#     {verification_results:[{result: rows}]}}`` shape the engine's
#     ``_execute_sql_query`` reads.
#   * ``_LitellmJudgeClient.llm.ainvoke(messages)`` folds the engine's
#     ``[SystemMessage, HumanMessage]`` into one prompt and calls the sync
#     ``_call_judge_with_fallback`` (litellm), returning an object with
#     ``.content`` — so a monkeypatched ``_call_judge`` / ``query_state`` still
#     intercepts every judge / SQL touchpoint.
#   * a sync driver (``asyncio.run`` from the sync ``main()``) invokes the async
#     engine; the verifier process has no running loop so this never collides.
#
# The module-level adapters below (``_execute_rubric_check_verifier`` /
# ``_run_tool_execution`` / ``_run_json_match`` / ``_run_response_check`` /
# ``run_benchmark_verifiers``) are thin glue: they resolve the SQL gym, build the
# in-sandbox ``model_response`` from the trajectory + final answer, and enrich the
# result for the run-detail panel — but ALL judging / SQL / JSON-diff / tool-call
# scoring runs through the verbatim ``VerifierEngine``.

import asyncio
import types as _types
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple


class _EngineLogger:
    """No-op stand-in for benchmark_runner's module ``logger`` (the verbatim
    engine logs verbosely; the FROZEN template has no logger of its own and must
    not spam the verifier stdout)."""

    def info(self, *a, **k): pass
    def warning(self, *a, **k): pass
    def error(self, *a, **k): pass
    def debug(self, *a, **k): pass


logger = _EngineLogger()


# ── langchain_core.messages shim ────────────────────────────────────────
# The verbatim engine does ``from langchain_core.messages import SystemMessage,
# HumanMessage`` inside ``_judge_rubric`` / ``_compare_with_llm``. In the sandbox
# there is no langchain; register a minimal module so the verbatim import
# resolves. The judge shim only reads ``.content`` off each message, so a real
# langchain (present in the test container) works too — we never shadow it.

class SystemMessage:
    def __init__(self, content):
        self.content = content


class HumanMessage:
    def __init__(self, content):
        self.content = content


class ToolMessage:
    def __init__(self, content, tool_call_id=None):
        self.content = content
        self.tool_call_id = tool_call_id


def _ensure_langchain_messages_shim() -> None:
    try:
        import langchain_core.messages  # noqa: F401  (real one wins in the test env)
        return
    except Exception:
        pass
    mod = _types.ModuleType("langchain_core.messages")
    mod.SystemMessage = SystemMessage
    mod.HumanMessage = HumanMessage
    mod.ToolMessage = ToolMessage
    pkg = sys.modules.get("langchain_core")
    if pkg is None:
        pkg = _types.ModuleType("langchain_core")
        pkg.__path__ = []  # mark as a package so submodule import resolves
        sys.modules["langchain_core"] = pkg
    pkg.messages = mod
    sys.modules["langchain_core.messages"] = mod


_ensure_langchain_messages_shim()


class _ProxyProtocolClient:
    """Sandbox adapter for the engine's ``protocol_client``. Satisfies the
    verbatim ``_execute_sql_query`` openenv branch by wrapping the synchronous
    ``query_state(gym, q)`` proxy call (PROXY_BASE ``/raw/<gym>/state``)."""

    def __init__(self, gym_name):
        self.gym_name = gym_name
        # The verbatim mcp-mode branch (never reached — execution_mode is always
        # "openenv") reads these; keep them present so the attribute access can't
        # AttributeError even on a defensive path.
        self.database_id = ""
        self.context = {}
        self.base_url = ""

    async def get_state(self, verify_queries=None):
        queries = verify_queries or []
        query = queries[0] if queries else ""
        try:
            rows = query_state(self.gym_name, query)
        except Exception as exc:  # noqa: BLE001 — surface as a failed state result
            return {"success": False, "error": str(exc)}
        return {"success": True, "data": {"verification_results": [{"result": rows}]}}


class _LitellmJudgeResponse:
    """The object the verbatim engine reads ``.content`` off (mirrors a
    LangChain ``AIMessage``)."""

    def __init__(self, content):
        self.content = content


class _LitellmJudge:
    """Stands in for ``LLMClient.llm`` — the engine calls ``.ainvoke(messages)``."""

    def __init__(self, judge_cfg):
        self._cfg = judge_cfg or {}

    async def ainvoke(self, messages):
        # Fold the engine's [SystemMessage, HumanMessage] list into the single
        # user prompt the sync ``_call_judge_with_fallback`` takes.
        parts: list[str] = []
        for m in messages or []:
            content = getattr(m, "content", None)
            if content is None:
                content = m if isinstance(m, str) else str(m)
            parts.append(content if isinstance(content, str) else str(content))
        prompt = "\n\n".join(p for p in parts if p)
        raw = _call_judge_with_fallback(
            self._cfg.get("model"),
            self._cfg.get("api_key_env"),
            float(self._cfg.get("temperature", 0.0) or 0.0),
            int(self._cfg.get("max_tokens", 2048) or 2048),
            prompt,
            judge=self._cfg,
        )
        return _LitellmJudgeResponse(raw)


class _LitellmJudgeClient:
    """Stands in for ``LLMClient`` — exposes ``.llm`` with the async ``ainvoke``
    the verbatim engine calls (``self.judge_llm_client.llm.ainvoke``)."""

    def __init__(self, judge_cfg):
        self.llm = _LitellmJudge(judge_cfg)


# ---------------------------------------------------------------------------
# Trace evidence for ``rubric_check`` (target="trace")
# ---------------------------------------------------------------------------
# Trace rubrics used to get ``{name, args}`` only — the judge saw the call but
# never its result, so any rubric about a *finding* was ungradeable. Results are
# included now, clipped per call so one schema dump can't eat the whole budget.
_TRACE_RESULT_MAX_CHARS = int(os.environ.get("RUBRIC_TRACE_RESULT_CHARS", "4000"))
_TRACE_RESULT_MIN_CHARS = int(os.environ.get("RUBRIC_TRACE_RESULT_MIN_CHARS", "200"))
_TRACE_EVIDENCE_MAX_CHARS = int(os.environ.get("RUBRIC_TRACE_EVIDENCE_CHARS", "200000"))


def _clip_json_value(value: Any, limit: int) -> str:
    """Compact-serialize *value* and clip it to *limit* chars, flagging the cut."""
    try:
        text = json.dumps(value, separators=(",", ":"), default=str)
    except (TypeError, ValueError):
        text = str(value)
    if len(text) <= limit:
        return text
    # Head+tail: results lead with schema boilerplate and carry the rows at the
    # end (``data_array``), so a head-only clip drops every value.
    head = limit // 3
    tail = limit - head
    return (
        text[:head]
        + f"… [{len(text) - limit} chars clipped] …"
        + text[-tail:]
    )


def _build_trace_evidence(
    tool_calls: List[Any],
    include_outcomes: bool = False,
) -> str:
    """Trace evidence: each call with its result, clipped to cap/n_calls."""
    calls = [tc for tc in tool_calls if isinstance(tc, dict)]
    per_call = _TRACE_EVIDENCE_MAX_CHARS // max(1, len(calls))
    per_call = max(_TRACE_RESULT_MIN_CHARS, min(_TRACE_RESULT_MAX_CHARS, per_call))
    entries: List[Dict[str, Any]] = []
    for tc in calls:
        entry: Dict[str, Any] = {
            "name": tc.get("name"),
            "args": tc.get("args") or tc.get("arguments"),
        }
        if include_outcomes:
            entry["accepted_by_gym"] = bool(tc.get("ok", True))
        result = tc.get("tool_execution_results")
        if result not in (None, {}, ""):
            entry["result"] = _clip_json_value(result, per_call)
        entries.append(entry)
    return json.dumps(entries, indent=2)


@dataclass
class VerifierConfig:
    """Configuration for a verifier.

    category:
      - "core"      -> must pass for the run to count as an overall success.
      - "secondary" -> contributes to the soft/weighted score and diagnostics,
                       but does NOT block overall success (e.g. process floors).
    weight: relative weight in the weighted soft score (default 1.0).
    """

    verifier_type: str
    validation_config: Dict[str, Any]
    name: Optional[str] = None
    description: Optional[str] = None
    gym_name: Optional[str] = None
    category: str = "core"
    weight: float = 1.0
def _extract_fenced_json_text(response_text: str) -> str:
    """Pull JSON body from a judge reply (handles ```json fences)."""
    text = (response_text or "").strip()
    if "```json" in text:
        return text.split("```json", 1)[1].split("```", 1)[0].strip()
    if "```" in text:
        return text.split("```", 1)[1].split("```", 1)[0].strip()
    return text


def _judge_json_parseable(response_text: str) -> bool:
    """True when the judge returned non-empty, JSON-parseable content."""
    text = (response_text or "").strip()
    if not text:
        return False
    try:
        json.loads(_extract_fenced_json_text(text))
        return True
    except Exception:
        return False


def _judge_rubric_response_usable(response_text: str) -> bool:
    """True when a rubric judge reply can be parsed (JSON or bare PASS/FAIL)."""
    if _judge_json_parseable(response_text):
        return True
    text = (response_text or "").strip()
    if not text:
        return False
    # A reply that STARTED the JSON contract but does not parse is truncated or
    # malformed, not a bare verdict -- report it as unusable so the retry loop
    # actually fires. Without this, a cut-off
    # ``{"verdict": "PASS", "motivation": "...`` counts as usable purely because
    # the token PASS appears in it, and ``_parse_rubric_verdict`` then resolves
    # the unparseable JSON to FAIL: a passing task scored as a failure, with the
    # truncated blob as its motivation. Longer evidence (a whole deliverable
    # file) makes the judge write longer replies, so this is reachable.
    if text.startswith("{") or text.startswith("```") or '"verdict"' in text:
        return False
    up = text.upper()
    return "PASS" in up or "FAIL" in up


def _spec_paths_under_input(spec) -> list:
    """Return every `arguments.path` in a file_check spec that points under the
    seeded input directory (or otherwise escapes the workspace root).

    A verifier must grade what the model PRODUCED. `SourceContext.resolve_path`
    (vendored, not ours to change) permits any workspace-relative path, so an
    input file is a legal source — and a sub-verifier aimed at one passes
    deterministically for a model that made zero tool calls.

    Normalization uses `posixpath.normpath`, which collapses `..` segments the
    same way `SourceContext.resolve_path`'s `Path.resolve()` does, so
    `output/../input/x.csv` is caught even though the literal string never
    starts with `input/`.

    The `input/` comparison is case-INSENSITIVE (lowercased after
    normalization): the runner's own dev platform (macOS) has a
    case-insensitive filesystem, so `Input/x.csv` / `INPUT/x.csv` resolve to
    the same seeded fixture as `input/x.csv` and must be flagged identically —
    otherwise a spec that validates clean locally scores 0.0 on a
    case-sensitive filesystem (cloud/ext4) purely because the guard let it
    through. This does not affect the traversal/workspace-escape handling
    below, which is case-agnostic already (`..` has no letters).

    Paths that net-normalize to a leading `..` (i.e. they climb above the
    workspace root, e.g. `../input/x.csv`) are flagged too, not left to
    `resolve_path`'s workspace-escape check. That check compares against the
    real runtime workspace directory, whose basename this harness assigns
    deterministically (`run_{run_number}`, see `prepare_workspace` in
    skill_tools.py) — so `../run_1/input/x.csv` climbs out and back into the
    *same* directory by name and legitimately resolves inside the workspace,
    silently defeating a check that only special-cases `input/`. Rejecting any
    spec path that leaves the workspace root at all closes that gap without
    needing to know the runtime directory name at spec-validation time, and
    costs nothing: a `file_check` spec has no legitimate reason to reference
    anything outside the workspace it's grading.
    """
    prefix = f"{INPUT_DIR_NAME}/"
    found = []

    def flag_if_offending(value: str) -> None:
        normalized = posixpath.normpath(value.replace("\\", "/"))
        normalized_ci = normalized.lower()
        escapes_workspace = normalized == ".." or normalized.startswith("../")
        if (
            normalized_ci == INPUT_DIR_NAME
            or normalized_ci.startswith(prefix)
            or escapes_workspace
        ):
            found.append(value)

    def walk(node):
        if isinstance(node, dict):
            for key, value in node.items():
                if key == "path":
                    if isinstance(value, str):
                        flag_if_offending(value)
                    elif isinstance(value, list):
                        # Fail closed: a list-valued `path` is unreachable
                        # today (every known spec shape uses a bare string),
                        # but a security-shaped guard must still inspect the
                        # strings inside rather than silently skip them.
                        for item in value:
                            if isinstance(item, str):
                                flag_if_offending(item)
                            else:
                                walk(item)
                    else:
                        walk(value)
                else:
                    walk(value)
        elif isinstance(node, list):
            for item in node:
                walk(item)

    walk(spec)
    return found


class VerifierEngine:
    """Runs database_state, response_check, tool_execution, json_match and
    rubric_check verifiers. Multi-gym aware."""

    def __init__(
        self,
        protocol_clients: Dict,
        llm_client: LLMClient,
        execution_mode: str = "openenv",
        judge_llm_client: Optional[LLMClient] = None,
    ):
        self.protocol_clients = protocol_clients
        self.llm_client = llm_client
        self.execution_mode = execution_mode
        # Dedicated LLM-as-judge for rubric_check / response_check grading. A
        # fixed strong model grades verdicts so they never depend on (or are
        # graded by) the candidate model under test. Falls back to the
        # candidate client (self-grading) when no independent judge is wired in,
        # so existing callers and tests keep working unchanged.
        self.judge_llm_client = judge_llm_client or llm_client

    async def _judge_ainvoke(self, messages: List[Any]) -> Any:
        """Invoke the judge via LLMClient.ainvoke when available (provider retry)."""
        client = self.judge_llm_client
        ainvoke = getattr(client, "ainvoke", None)
        if callable(ainvoke):
            return await ainvoke(messages)
        return await client.llm.ainvoke(messages)

    async def _invoke_judge_llm(
        self,
        messages: List[Any],
        *,
        response_usable,
    ) -> str:
        """Call the judge LLM, retrying on empty or unparseable JSON replies.

        Provider transport/rate-limit retries (#104) run inside
        ``judge_llm_client.ainvoke``; this outer loop only handles content
        flakes (empty body or JSON that fails to parse).
        """
        max_attempts = max(1, _env_int("TH_JUDGE_JSON_RETRY_MAX", 3))
        delay = 0.5
        last_text = ""
        for attempt in range(1, max_attempts + 1):
            response = await self._judge_ainvoke(messages)
            last_text = _normalize_ai_content(response)
            if response_usable(last_text):
                return last_text
            if attempt < max_attempts:
                logger.warning(
                    "Judge returned empty/unparseable JSON (attempt %d/%d); "
                    "retrying in %.1fs",
                    attempt,
                    max_attempts,
                    delay,
                )
                await asyncio.sleep(delay)
                delay = min(delay * 2, 4.0)
        return last_text

    def _get_protocol_client_for_gym(self, gym_name: Optional[str] = None):
        """Get the correct protocol client for a specific gym"""
        if not gym_name:

            return list(self.protocol_clients.values())[0]

        if gym_name in self.protocol_clients:
            return self.protocol_clients[gym_name]

        return list(self.protocol_clients.values())[0]

    async def execute_verifier(
        self,
        verifier: VerifierConfig,
        model_response: Dict[str, Any],
        gym_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Execute a single verifier"""
        logger.info(f"Executing verifier: {verifier.verifier_type}")

        if verifier.verifier_type == "database_state":
            return await self._execute_database_state_verifier(
                verifier.validation_config, gym_name
            )
        elif verifier.verifier_type == "response_check":
            return await self._execute_response_check_verifier(
                verifier.validation_config,
                model_response,
                gym_name,
            )
        elif verifier.verifier_type == "tool_execution":
            return await self._execute_tool_execution_verifier(
                verifier.validation_config, model_response
            )
        elif verifier.verifier_type == "json_match":
            return self._execute_json_match_verifier(
                verifier.validation_config, model_response
            )
        elif verifier.verifier_type == "source_json_match":
            return await self._execute_source_json_match_verifier(
                verifier.validation_config, model_response
            )
        elif verifier.verifier_type == "rubric_check":
            return await self._execute_rubric_check_verifier(
                verifier.validation_config, model_response, gym_name
            )
        elif verifier.verifier_type == "file_check":
            return await self._execute_file_check_verifier(
                verifier.validation_config, model_response, gym_name
            )
        elif verifier.verifier_type == "gym_state_check":
            return await self._execute_gym_state_check_verifier(
                verifier.validation_config, gym_name
            )
        else:
            return {
                "passed": False,
                "error": f"Unsupported verifier type: {verifier.verifier_type}",
            }

    async def _execute_file_check_verifier(
        self,
        validation_config: Dict[str, Any],
        model_response: Dict[str, Any],
        gym_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Grade the files the model produced in its per-run skills workspace.

        The skills flow runs the model's `bash` tool inside a fresh per-run
        workspace and records its path as `model_response["workspace_dir"]`.
        This verifier points rl-world-verifiers straight at that directory and
        evaluates an embedded spec (deterministic file-content checks and/or
        LLM rubrics). The scalar reward (0..1) maps to pass/fail via
        `pass_threshold` (default 1.0), so it flows through the existing
        core/secondary scoring path like any other verifier.
        """
        spec = validation_config.get("verifier_spec")
        if not spec:
            return {"passed": False, "error": "file_check: missing 'verifier_spec'"}

        offending = _spec_paths_under_input(spec)
        if offending:
            return {
                "passed": False,
                "error": (
                    "file_check: verifier spec reads seeded input files "
                    f"({', '.join(offending)}). A verifier must grade what the model "
                    f"produced, not the fixtures it was given. Point these at the "
                    f"deliverable paths instead."
                ),
            }

        threshold = float(validation_config.get("pass_threshold", 1.0))
        # Clamp to (0, 1]: a reward of 0.0 (no file produced, or every check failed)
        # must never satisfy the pass test, so pass_threshold <= 0 is not allowed to
        # grant a pass, and a threshold above 1.0 is capped so a perfect run passes.
        threshold = min(max(threshold, 1e-9), 1.0)

        workspace_dir = (model_response or {}).get("workspace_dir")
        if not workspace_dir or not os.path.isdir(workspace_dir):
            return {
                "passed": False,
                "error": "file_check: no workspace to grade. A file_check verifier "
                "auto-enables skills, so this only happens if TH_ENABLE_SKILLS was "
                "forced off (=0). Unset it (or set 1) so the bash tool writes a per-run "
                "workspace.",
                "gym_name": gym_name,
            }

        try:
            from rl_world_verifiers import run_verifier
        except Exception as exc:  # pragma: no cover - dependency/setup issue
            return {
                "passed": False,
                "error": f"file_check: rl_world_verifiers import failed ({exc}). "
                "Install the skills dependencies with `pip install -r requirements.txt` "
                "and run under Python >= 3.10 "
                "(see README.md, 'Document skills + file_check verifier').",
            }

        seeded_paths = (model_response or {}).get("seeded_input_paths") or []
        produced = list_workspace_files(workspace_dir, seeded_paths)
        produced_names = [f.get("path") if isinstance(f, dict) else f for f in produced]

        out_dir = tempfile.mkdtemp(prefix="filecheck_out_")
        try:
            spec_path = os.path.join(out_dir, "verifier.json")
            with open(spec_path, "w", encoding="utf-8") as fh:
                json.dump(spec, fh)

            # run_verifier is synchronous; keep the event loop free.
            payload = await asyncio.to_thread(
                run_verifier,
                spec_path=Path(spec_path),
                workspace_dir=Path(workspace_dir),
                verifier_dir=Path(out_dir),
            )
            reward = float(payload.get("reward", 0.0))
            summary = payload.get("summary", {})
            return {
                "passed": reward >= threshold,
                "reward": reward,
                "threshold": threshold,
                "workspace_files": produced_names,
                "summary": summary,
                "assertion_results": payload.get("assertion_results"),
                "gym_name": gym_name,
            }
        except Exception as exc:
            return {
                "passed": False,
                "error": f"file_check: verifier run failed: {exc}",
                "workspace_files": produced_names,
                "gym_name": gym_name,
            }
        finally:
            shutil.rmtree(out_dir, ignore_errors=True)

    async def _execute_database_state_verifier(
        self,
        validation_config: Dict[str, Any],
        gym_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Execute database state verifier"""
        sql_query = validation_config.get("sql_query")
        expected_value = validation_config.get("expected_value")
        comparison_type = validation_config.get("comparison_type", "equals")

        if not sql_query:
            return {"passed": False, "error": "No SQL query provided"}

        logger.info(f"Executing SQL query: {sql_query}")
        if gym_name:
            logger.info(f"  Target gym: {gym_name}")

        result = await self._execute_sql_query(sql_query, gym_name)

        if not result["success"]:
            return {
                "passed": False,
                "error": f"SQL query failed: {result.get('error')}",
                "query": sql_query,
            }

        actual_value = self._extract_value_from_sql_result(result)

        logger.info(f"SQL result - Expected: {expected_value}, Actual: {actual_value}")

        comparison_result = self._compare_values(
            actual_value, expected_value, comparison_type
        )

        return {
            "passed": comparison_result["passed"],
            "expected": expected_value,
            "actual": actual_value,
            "comparison_type": comparison_type,
            "query": sql_query,
            "gym_name": gym_name,
            "details": comparison_result.get("details"),
        }

    def _execute_json_match_verifier(
        self,
        validation_config: Dict[str, Any],
        model_response: Dict[str, Any],
    ) -> Dict[str, Any]:
        """Compare parsed JSON in the model response to expected values."""
        expected = validation_config.get("expected") or {}
        if not expected:
            return {"passed": False, "error": "json_match: missing expected object"}

        tolerance = float(validation_config.get("float_tolerance", 0.02))
        text = self._extract_llm_content(model_response)
        parsed = _extract_json_from_text(text)
        if parsed is None:
            return {
                "passed": False,
                "error": "json_match: could not parse JSON from model response",
                "response_preview": text[:500],
            }

        mismatches = self._diff_json_expected(parsed, expected, tolerance)
        return {
            "passed": not mismatches,
            "expected": expected,
            "actual": parsed,
            "mismatches": mismatches,
            "float_tolerance": tolerance,
        }

    def _diff_json_expected(
        self,
        actual: Dict[str, Any],
        expected: Dict[str, Any],
        tolerance: float,
        prefix: str = "",
    ) -> List[str]:
        mismatches: List[str] = []
        for key, exp in expected.items():
            path = f"{prefix}{key}" if not prefix else f"{prefix}.{key}"
            if key not in actual:
                mismatches.append(f"{path}: missing in response")
                continue
            act = actual[key]
            if isinstance(exp, bool):
                if act is not exp and str(act).lower() != str(exp).lower():
                    mismatches.append(f"{path}: expected {exp!r}, got {act!r}")
            elif isinstance(exp, (int, float)):
                if act is None:
                    mismatches.append(f"{path}: expected {exp!r}, got null")
                    continue
                try:
                    act_num = float(act)
                except (TypeError, ValueError):
                    mismatches.append(f"{path}: expected {exp!r}, got {act!r}")
                else:
                    if abs(act_num - float(exp)) > tolerance:
                        mismatches.append(
                            f"{path}: expected {exp}, got {act} (tol {tolerance})"
                        )
            elif isinstance(exp, str):
                if str(act).strip() != exp:
                    mismatches.append(f"{path}: expected {exp!r}, got {act!r}")
            elif isinstance(exp, dict) and isinstance(act, dict):
                mismatches.extend(
                    self._diff_json_expected(act, exp, tolerance, path)
                )
            elif exp != act:
                mismatches.append(f"{path}: expected {exp!r}, got {act!r}")
        return mismatches

    async def _execute_source_json_match_verifier(
        self,
        validation_config: Dict[str, Any],
        model_response: Dict[str, Any],
    ) -> Dict[str, Any]:
        """Compare final JSON paths to values computed from task data sources."""
        matches = validation_config.get("matches") or []
        if not matches:
            return {"passed": False, "error": "source_json_match: missing matches[]"}

        text = self._extract_llm_content(model_response)
        parsed = _extract_json_from_text(text)
        if parsed is None:
            return {
                "passed": False,
                "error": "source_json_match: could not parse JSON from model response",
                "response_preview": text[:500],
            }

        source_configs = self._source_json_match_sources(validation_config)
        default_tolerance = float(validation_config.get("float_tolerance", 0.02))
        source_cache: Dict[str, Dict[str, Any]] = {}
        mismatches: List[str] = []
        evaluations: List[Dict[str, Any]] = []

        for idx, match in enumerate(matches):
            if not isinstance(match, dict):
                mismatches.append(f"matches[{idx}]: expected object, got {type(match).__name__}")
                continue

            source_name = match.get("source_ref") or match.get("source")
            answer_path = match.get("answer_path")
            expected_path = match.get("expected_path")
            comparison_type = match.get("comparison_type", "equals")
            tolerance = float(match.get("float_tolerance", default_tolerance))

            if not source_name:
                mismatches.append(f"matches[{idx}]: missing source_ref/source")
                continue
            if not answer_path or not expected_path:
                mismatches.append(f"matches[{idx}]: missing answer_path or expected_path")
                continue

            if source_name not in source_configs:
                mismatches.append(f"{answer_path}: unknown source_ref/source {source_name!r}")
                continue

            if source_name not in source_cache:
                source_result = await self._run_source_json_match_source(
                    source_name, source_configs[source_name]
                )
                source_cache[source_name] = source_result
            else:
                source_result = source_cache[source_name]

            if not source_result.get("success"):
                mismatches.append(f"{answer_path}: source {source_name!r} failed: {source_result.get('error')}")
                evaluations.append(
                    {
                        "source": source_name,
                        "answer_path": answer_path,
                        "expected_path": expected_path,
                        "error": source_result.get("error"),
                    }
                )
                continue

            try:
                answer_found, actual_value = _get_simple_json_path(parsed, answer_path)
                expected_found, expected_value = _get_simple_json_path(
                    source_result.get("data"), expected_path
                )
            except ValueError as exc:
                mismatches.append(f"{answer_path}: invalid JSON path: {exc}")
                continue

            if not answer_found:
                mismatches.append(f"{answer_path}: missing in response")
                continue
            if not expected_found:
                mismatches.append(
                    f"{answer_path}: expected path {expected_path!r} missing in source {source_name!r}"
                )
                continue

            comparison = self._compare_source_json_values(
                actual_value, expected_value, comparison_type, tolerance
            )
            evaluations.append(
                {
                    "source": source_name,
                    "answer_path": answer_path,
                    "expected_path": expected_path,
                    "comparison_type": comparison_type,
                    "float_tolerance": tolerance,
                    "expected": expected_value,
                    "actual": actual_value,
                    "passed": comparison["passed"],
                    "details": comparison.get("details"),
                }
            )
            if not comparison["passed"]:
                mismatches.append(
                    f"{answer_path}: expected {expected_value!r} from {source_name}{expected_path}, "
                    f"got {actual_value!r} ({comparison.get('details')})"
                )

        return {
            "passed": not mismatches,
            "actual": parsed,
            "mismatches": mismatches,
            "evaluations": evaluations,
            "source_results": {
                name: {
                    "gym_name": value.get("gym_name"),
                    "query": value.get("query"),
                    "data": value.get("data"),
                    "error": value.get("error"),
                }
                for name, value in source_cache.items()
            },
        }

    def _source_json_match_sources(
        self, validation_config: Dict[str, Any]
    ) -> Dict[str, Dict[str, Any]]:
        sources: Dict[str, Dict[str, Any]] = {}

        for ref in validation_config.get("data_references") or []:
            if not isinstance(ref, dict):
                continue
            ref_id = ref.get("id")
            verifier = ref.get("verifier") or {}
            if ref_id and isinstance(verifier, dict):
                sources[str(ref_id)] = {
                    "gym_name": verifier.get("gym_name") or verifier.get("target_gym_server"),
                    "sql_query": verifier.get("sql_query"),
                    "label": ref.get("label"),
                    "browser_path": ref.get("browser_path") or ref.get("path"),
                }

        for source in validation_config.get("sources") or []:
            if not isinstance(source, dict):
                continue
            name = source.get("name") or source.get("id")
            if not name:
                continue
            sources[str(name)] = {
                "gym_name": source.get("gym_name") or source.get("target_gym_server"),
                "sql_query": source.get("sql_query"),
                "label": source.get("label"),
                "browser_path": source.get("browser_path") or source.get("path"),
            }

        return sources

    async def _run_source_json_match_source(
        self, source_name: str, source_config: Dict[str, Any]
    ) -> Dict[str, Any]:
        query = source_config.get("sql_query")
        gym_name = source_config.get("gym_name")
        if not query:
            return {
                "success": False,
                "error": f"source_json_match source {source_name!r} missing sql_query",
                "gym_name": gym_name,
            }

        result = await self._execute_sql_query(query, gym_name)
        if not result.get("success"):
            return {
                "success": False,
                "error": result.get("error", "query failed"),
                "gym_name": gym_name,
                "query": query,
            }

        return {
            "success": True,
            "gym_name": gym_name,
            "query": query,
            "data": self._extract_data_from_sql_result(result),
        }

    def _extract_data_from_sql_result(self, result: dict) -> Any:
        """Normalize SQL results for source_json_match expected-path lookups."""
        if not result:
            return None

        result_data = result.get("result", {})
        if isinstance(result_data, dict):
            for key in ("data", "rows", "result"):
                if key in result_data:
                    result_data = result_data[key]
                    break

        if isinstance(result_data, list):
            if len(result_data) == 1:
                return result_data[0]
            return result_data

        return result_data

    def _compare_source_json_values(
        self,
        actual: Any,
        expected: Any,
        comparison_type: str,
        tolerance: float,
    ) -> Dict[str, Any]:
        """Compare source-backed values with JSON-match-style coercion."""
        try:
            if isinstance(expected, bool):
                actual_bool = actual if isinstance(actual, bool) else str(actual).lower()
                expected_bool = expected if isinstance(actual, bool) else str(expected).lower()
                passed = actual_bool == expected_bool
            elif isinstance(expected, (int, float)) and not isinstance(expected, bool):
                actual_num = float(actual)
                expected_num = float(expected)
                if comparison_type in ("equals", "approx_equals"):
                    passed = abs(actual_num - expected_num) <= tolerance
                elif comparison_type == "not_equals":
                    passed = abs(actual_num - expected_num) > tolerance
                elif comparison_type == "greater_than":
                    passed = actual_num > expected_num
                elif comparison_type == "greater_or_equal":
                    passed = actual_num >= expected_num
                elif comparison_type == "less_than":
                    passed = actual_num < expected_num
                elif comparison_type == "less_or_equal":
                    passed = actual_num <= expected_num
                else:
                    return {
                        "passed": False,
                        "details": f"Unknown comparison type: {comparison_type}",
                    }
                return {
                    "passed": passed,
                    "details": f"Comparison {comparison_type}: {actual_num} vs {expected_num} (tol {tolerance})",
                }
            elif isinstance(expected, str):
                actual_str = str(actual).strip()
                if comparison_type == "equals":
                    passed = actual_str == expected
                elif comparison_type == "not_equals":
                    passed = actual_str != expected
                elif comparison_type == "contains":
                    passed = expected in actual_str
                else:
                    return {
                        "passed": False,
                        "details": f"Unknown comparison type for string: {comparison_type}",
                    }
            else:
                if comparison_type == "equals":
                    passed = actual == expected
                elif comparison_type == "not_equals":
                    passed = actual != expected
                else:
                    return {
                        "passed": False,
                        "details": f"Unknown comparison type for value: {comparison_type}",
                    }

            return {
                "passed": passed,
                "details": f"Comparison {comparison_type}: {actual!r} vs {expected!r}",
            }
        except Exception as e:
            return {"passed": False, "details": f"Comparison error: {e}"}

    async def _execute_response_check_verifier(
        self,
        validation_config: Dict[str, Any],
        model_response: Dict[str, Any],
        gym_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Execute response check verifier using LLM-as-judge"""
        sql_query = validation_config.get("sql_query")
        comparison_prompt = validation_config.get("comparison_prompt")
        minimum_comparison_value = validation_config.get("minimum_comparison_value", 7)

        if not comparison_prompt:
            comparison_prompt = (
                "Does the AI assistant response accurately reflect the database result above "
                "and confirm that the requested action was completed?"
            )

        if not sql_query:
            return {"passed": False, "error": "Missing sql_query"}

        sql_result = await self._execute_sql_query(sql_query, gym_name)

        if not sql_result["success"]:
            return {
                "passed": False,
                "error": f"SQL query failed: {sql_result.get('error')}",
            }

        llm_response_text = self._extract_llm_content(model_response)

        judge_result = await self._compare_with_llm(
            sql_result, llm_response_text, comparison_prompt, minimum_comparison_value
        )

        return judge_result

    # ------------------------------------------------------------------
    # Deliverable-file evidence for ``rubric_check`` (target="workspace_files")
    #
    # The judge's other targets read the agent's *talk* (final message) or its
    # *actions* (tool trace). Neither proves anything about the files the task
    # actually asked for: an agent can narrate a correct answer in chat and
    # write nothing, and score full marks on a final_answer rubric. This target
    # puts the produced files themselves in front of the judge, so a rubric
    # about the deliverable is judged against the deliverable.
    #
    # Deterministic per-figure assertions still belong in ``file_check``; this
    # is for the claims that need reading -- "the write-up explains why X rather
    # than Y", "the recommendation follows from the figures it reports".
    # ------------------------------------------------------------------
    _WORKSPACE_EVIDENCE_PER_FILE_CHARS = 20000
    _WORKSPACE_EVIDENCE_TOTAL_CHARS = 60000

    @staticmethod
    def _render_workspace_file(abs_path: str) -> str:
        """Render one deliverable as judge-readable text.

        Spreadsheets are dumped cell by cell (``Sheet!B5 = 1234``) because the
        judge cannot read the binary and the cell address is what the
        instruction pins. Everything else is read as text.
        """
        if os.path.splitext(abs_path)[1].lower() == ".xlsx":
            try:
                from openpyxl import load_workbook

                workbook = load_workbook(abs_path, data_only=True)
            except Exception as exc:  # unreadable workbook -> say so, fail closed
                return f"<unreadable workbook: {exc}>"
            lines = [
                f"{sheet.title}!{cell.coordinate} = {cell.value}"
                for sheet in workbook.worksheets
                for row in sheet.iter_rows()
                for cell in row
                if cell.value is not None
            ]
            return "\n".join(lines) or "<workbook has no non-empty cells>"
        try:
            with open(abs_path, "r", encoding="utf-8", errors="replace") as handle:
                return handle.read()
        except Exception as exc:
            return f"<unreadable file: {exc}>"

    def _build_workspace_file_evidence(
        self, model_response: Dict[str, Any], files: List[str]
    ) -> str:
        """Concatenate the declared deliverables into one evidence block.

        Fails CLOSED at every step: no workspace, no declared files, a path that
        escapes the workspace, or a file the agent never wrote all render as an
        explicit marker rather than silently shrinking the evidence, so the
        judge sees the absence instead of judging a blank.
        """
        workspace = (model_response or {}).get("workspace_dir")
        if not workspace or not os.path.isdir(str(workspace)):
            return (
                "<no workspace directory available: the deliverable files could "
                "not be read, so nothing about them is established>"
            )
        if not files:
            return (
                "<this rubric declared no 'files': nothing about the deliverables "
                "is established>"
            )

        workspace_root = os.path.realpath(str(workspace))
        chunks: List[str] = []
        budget = self._WORKSPACE_EVIDENCE_TOTAL_CHARS
        for rel_path in files:
            rel = str(rel_path)
            abs_path = os.path.realpath(os.path.join(workspace_root, rel))
            if not (abs_path == workspace_root or abs_path.startswith(workspace_root + os.sep)):
                chunks.append(f"--- {rel} ---\n<rejected: path escapes the workspace>")
                continue
            if not os.path.isfile(abs_path):
                chunks.append(f"--- {rel} ---\nMISSING (the agent never wrote this file)")
                continue
            body = self._render_workspace_file(abs_path)
            if len(body) > self._WORKSPACE_EVIDENCE_PER_FILE_CHARS:
                body = body[: self._WORKSPACE_EVIDENCE_PER_FILE_CHARS] + "\n<clipped>"
            if budget <= 0:
                chunks.append(f"--- {rel} ---\n<omitted: evidence budget exhausted>")
                continue
            body = body[:budget]
            budget -= len(body)
            chunks.append(f"--- {rel} ---\n{body}")
        return "\n\n".join(chunks)

    async def _execute_rubric_check_verifier(
        self,
        validation_config: Dict[str, Any],
        model_response: Dict[str, Any],
        gym_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """LLM-as-judge against a single ATOMIC rubric (one yes/no claim).

        Realises the rubric-style verifier: the judge is given exactly one
        rubric, the relevant slice of the model's behaviour, and an optional
        verification-only reference, and returns PASS/FAIL plus a one-line
        motivation. The result carries category/weight like any other verifier,
        so a binary task fails if any core rubric fails, while a scored task can
        sum the weights of the passed rubrics.

        validation_config keys:
          rubric    (str, required) one atomic yes/no claim to evaluate.
          target    (str) which slice to judge: "final_answer" (default),
                    "final_state", "trace", or "workspace_files".
          files     (list[str], required for target="workspace_files") the
                    workspace-relative deliverable paths to put in front of the
                    judge, e.g. ["metrics.json", "report.md"].
          sql_query (str, optional) authoritative state context (final_state).
          expected  (any, optional) verification-only reference; never shown to
                    the model, only to the judge.
        """
        rubric = (validation_config.get("rubric") or "").strip()
        if not rubric:
            return {"passed": False, "error": "rubric_check: missing 'rubric'"}

        target = (validation_config.get("target") or "final_answer").lower()
        reference = validation_config.get("expected")

        if target == "trace":
            tool_calls = model_response.get("tool_calls") or []
            # Opt-in: flags calls the gym rejected. Kept opt-in because it can
            # only turn passes into fails on existing rubrics.
            include_outcomes = bool(validation_config.get("include_tool_outcomes", False))
            evidence_label = (
                "Tool-call trace (each entry is one call the agent made and the "
                "result it received back, clipped to a per-call budget; a missing "
                "'result' means the call returned nothing)"
            )
            if include_outcomes:
                evidence_label += (
                    " — accepted_by_gym=false means the gym REJECTED the call and "
                    "nothing changed; such a call proves nothing"
                )
            evidence = _build_trace_evidence(tool_calls, include_outcomes)
        elif target == "workspace_files":
            evidence_label = (
                "Deliverable files the agent wrote to its workspace (each file "
                "appears under a '--- <path> ---' header; a file shown as MISSING "
                "was never produced and proves nothing)"
            )
            evidence = self._build_workspace_file_evidence(
                model_response, validation_config.get("files") or []
            )
        else:
            evidence_label = "AI assistant final answer"
            evidence = self._extract_llm_content(model_response)

        sql_context = None
        sql_query = validation_config.get("sql_query")
        if sql_query:
            sql_res = await self._execute_sql_query(sql_query, gym_name)
            if not sql_res.get("success"):
                return {
                    "passed": False,
                    "error": f"rubric_check: context SQL failed: {sql_res.get('error')}",
                    "rubric": rubric,
                }
            sql_context = json.dumps(sql_res.get("result", {}), indent=2)

        return await self._judge_rubric(
            rubric, target, evidence_label, evidence, reference, sql_context
        )

    async def _judge_rubric(
        self,
        rubric: str,
        target: str,
        evidence_label: str,
        evidence: str,
        reference: Any = None,
        sql_context: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Ask the judge LLM whether one atomic rubric holds. Strict PASS/FAIL."""
        from langchain_core.messages import SystemMessage, HumanMessage

        system_prompt = (
            "You are a strict evaluation judge. You are given ONE atomic rubric: "
            "a single yes/no statement about an AI assistant's behaviour. Decide "
            "whether the rubric is satisfied by the evidence below. Judge ONLY "
            "this one rubric and nothing else. If the evidence is insufficient to "
            "satisfy the rubric, the verdict is FAIL.\n\n"
            "Respond with ONLY a JSON object in this format:\n"
            "{\n"
            '  "verdict": "PASS" | "FAIL",\n'
            '  "motivation": "<one sentence naming the deviation, or why it passes>"\n'
            "}"
        )

        parts = [f"Rubric (atomic — judge on its own):\n{rubric}\n"]
        if reference not in (None, {}, "", []):
            parts.append(
                "Reference (verification-only ground truth — do NOT treat as the "
                f"assistant's output):\n{json.dumps(reference, indent=2)}\n"
            )
        if sql_context is not None:
            parts.append(f"Authoritative state (from the data warehouse):\n{sql_context}\n")
        parts.append(f"{evidence_label}:\n{evidence}\n")
        parts.append("Return your verdict as JSON.")
        user_prompt = "\n".join(parts)

        messages = [
            SystemMessage(content=system_prompt),
            HumanMessage(content=user_prompt),
        ]

        try:
            response_text = await self._invoke_judge_llm(
                messages, response_usable=_judge_rubric_response_usable
            )
            if not _judge_rubric_response_usable(response_text):
                return {
                    "passed": False,
                    "error": (
                        "rubric_check judge failed: empty or unparseable JSON "
                        f"after {max(1, _env_int('TH_JUDGE_JSON_RETRY_MAX', 3))} attempt(s)"
                    ),
                    "rubric": rubric,
                }
            verdict, motivation = self._parse_rubric_verdict(response_text)
            return {
                "passed": verdict == "PASS",
                "verdict": verdict,
                "rubric": rubric,
                "target": target,
                "motivation": motivation,
            }
        except Exception as e:
            logger.error(f"rubric_check judge failed: {e}")
            return {
                "passed": False,
                "error": f"rubric_check judge failed: {e}",
                "rubric": rubric,
            }

    @staticmethod
    def _parse_rubric_verdict(response_text: str) -> Tuple[str, str]:
        """Parse a judge reply into (verdict, motivation).

        Robust to ```json fences and to a bare PASS/FAIL. Fail-closed: any
        ambiguity (no clear verdict, malformed JSON, or both tokens present)
        resolves to FAIL so a flaky judge can never silently pass a task.
        """
        text = (response_text or "").strip()
        fenced = text
        if "```json" in fenced:
            fenced = fenced.split("```json", 1)[1].split("```", 1)[0].strip()
        elif "```" in fenced:
            fenced = fenced.split("```", 1)[1].split("```", 1)[0].strip()

        verdict: Optional[str] = None
        motivation = ""
        parsed_json = False
        try:
            obj = json.loads(fenced)
            parsed_json = True
            v = str(obj.get("verdict", "")).strip().upper()
            if v in ("PASS", "FAIL"):
                verdict = v
            motivation = str(obj.get("motivation", "")).strip()
        except Exception:
            pass

        if verdict is None:
            # The judge answered in the JSON contract but its verdict field was
            # missing/invalid: that is ambiguity, so fail closed. Only fall back to
            # the bare-token scan when the reply was not JSON at all -- otherwise a
            # motivation like "the model did not PASS" would score as a PASS.
            if parsed_json:
                verdict = "FAIL"
            else:
                up = text.upper()
                if "PASS" in up and "FAIL" not in up:
                    verdict = "PASS"
                else:
                    verdict = "FAIL"
            if not motivation:
                motivation = text[:200]
        return verdict, motivation

    async def _execute_gym_state_check_verifier(
        self,
        validation_config: Dict[str, Any],
        gym_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Grade the gym's FINAL state, for gyms with no SQL surface.

        The trace-based verifiers record what the model *asked for*, not what the
        gym *did*: a rejected call, a call against a fabricated node id, a fill
        applied to the wrong node, or a create-then-delete all leave a "correct"
        trace behind. This verifier reads the gym's authoritative post-episode
        state (GET /state?include_db=true -> "database_state") and asserts on it.

        Config:
          find           {key: value, ...}  select objects anywhere in the state
                                            tree whose keys all match exactly.
          expect_matches int                required number of matching objects
                                            (1 rejects both absent and duplicate).
          expect         {dotted.path: value}  asserted on the single match.
                                            List indices are numeric segments,
                                            e.g. "fills.0.color.r".
          float_tolerance float             numeric comparison tolerance.
        """
        find = validation_config.get("find") or {}
        expect = validation_config.get("expect") or {}
        expect_matches = validation_config.get("expect_matches")
        tolerance = float(validation_config.get("float_tolerance", 0.02))

        if not find:
            return {"passed": False, "error": "gym_state_check: missing 'find' selector"}

        if self.execution_mode != "openenv":
            return {
                "passed": False,
                "error": "gym_state_check: requires execution_mode 'openenv'",
            }

        protocol_client = self._get_protocol_client_for_gym(gym_name)
        state_result = await protocol_client.get_state(include_db=True)
        if not state_result.get("success"):
            return {
                "passed": False,
                "error": f"gym_state_check: /state failed: {state_result.get('error')}",
            }

        tree = (state_result.get("data") or {}).get("database_state")
        if tree is None:
            return {
                "passed": False,
                "error": "gym_state_check: /state returned no database_state "
                "(gym never materialised a tree for this episode)",
            }

        matches = [
            obj
            for obj in _walk_json_objects(tree)
            if all(k in obj and obj[k] == v for k, v in find.items())
        ]

        if expect_matches is not None and len(matches) != int(expect_matches):
            return {
                "passed": False,
                "error": (
                    f"gym_state_check: expected {expect_matches} object(s) matching "
                    f"{find}, found {len(matches)} in the gym's final state"
                ),
                "find": find,
                "match_count": len(matches),
            }
        if not matches:
            return {
                "passed": False,
                "error": f"gym_state_check: no object matching {find} in final state",
                "find": find,
                "match_count": 0,
            }

        target = matches[0]
        mismatches: List[str] = []
        for path, exp in expect.items():
            found, act = _resolve_dotted_path(target, path)
            if not found:
                mismatches.append(f"{path}: missing in final state")
            elif isinstance(exp, bool) or isinstance(act, bool):
                if act is not exp:
                    mismatches.append(f"{path}: expected {exp!r}, got {act!r}")
            elif isinstance(exp, (int, float)) and isinstance(act, (int, float)):
                if abs(float(act) - float(exp)) > tolerance:
                    mismatches.append(f"{path}: expected {exp!r}, got {act!r}")
            elif act != exp:
                mismatches.append(f"{path}: expected {exp!r}, got {act!r}")

        return {
            "passed": not mismatches,
            "find": find,
            "match_count": len(matches),
            "mismatches": mismatches,
            "float_tolerance": tolerance,
        }

    async def _execute_tool_execution_verifier(
        self, validation_config: Dict[str, Any], model_response: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Execute tool execution verifier"""
        expected_tools = validation_config.get("expected_tools", [])
        minimum_tool_calls = validation_config.get("minimum_tool_calls", 1)
        tool_name = validation_config.get("tool_name", "")
        occurrence_expected_value = validation_config.get("occurrence_expected_value")
        occurrence_comparison_type = validation_config.get(
            "occurrence_comparison_type", "equals"
        )
        # A process floor asks "did the model really do the work?". A call the gym
        # REJECTED is not work: the model can satisfy the floor by naming the tool
        # with bogus arguments. Opt-in (default False) so existing task grades are
        # unchanged; new tasks should set it true.
        exclude_errored_calls = bool(validation_config.get("exclude_errored_calls", False))
        # Call bounds. min_calls/max_calls bound the calls made to the tools in
        # expected_tools ("open each of the 72 messages exactly once"), while
        # min_total_tool_calls bounds the whole run's call count (a synonym of
        # minimum_tool_calls; the stricter of the two wins).
        min_calls = validation_config.get("min_calls")
        max_calls = validation_config.get("max_calls")
        min_total_tool_calls = validation_config.get("min_total_tool_calls")
        if min_total_tool_calls is not None:
            minimum_tool_calls = max(int(minimum_tool_calls), int(min_total_tool_calls))

        all_calls = model_response.get("tool_calls") or []
        if exclude_errored_calls:
            counted = [tc for tc in all_calls if tc.get("ok", True)]
        else:
            counted = list(all_calls)
        tools_called = [tc["name"] for tc in counted]
        errored_calls = [tc["name"] for tc in all_calls if not tc.get("ok", True)]

        logger.info(f"Expected tools: {expected_tools}, Called: {tools_called}")
        if errored_calls:
            logger.info(
                f"Gym-rejected calls: {errored_calls} "
                f"({'excluded from' if exclude_errored_calls else 'COUNTED toward'} the floor)"
            )

        missing_tools = [tool for tool in expected_tools if tool not in tools_called]

        passed = not missing_tools and len(tools_called) >= minimum_tool_calls

        expected_tool_calls = sum(1 for t in tools_called if t in set(expected_tools))
        call_bounds = None
        if min_calls is not None or max_calls is not None:
            call_bounds = {
                "expected_tool_calls": expected_tool_calls,
                "min_calls": min_calls,
                "max_calls": max_calls,
            }
            if min_calls is not None and expected_tool_calls < int(min_calls):
                passed = False
            if max_calls is not None and expected_tool_calls > int(max_calls):
                passed = False

        occurrence_details = None
        if tool_name:
            tool_occurrences = sum(1 for t in tools_called if t == tool_name)
            occurrence_details = {
                "tool_name": tool_name,
                "expected": occurrence_expected_value,
                "actual": tool_occurrences,
                "comparison_type": occurrence_comparison_type,
            }

            if occurrence_expected_value is not None:
                try:
                    expected_count = int(occurrence_expected_value)
                except (TypeError, ValueError):
                    return {
                        "passed": False,
                        "error": f"Invalid occurrence_expected_value: {occurrence_expected_value}",
                        "tools_called": tools_called,
                        "missing_tools": missing_tools,
                    }

                if occurrence_comparison_type == "equals":
                    occurrence_passed = tool_occurrences == expected_count
                elif occurrence_comparison_type == "greater_than":
                    occurrence_passed = tool_occurrences > expected_count
                elif occurrence_comparison_type == "greater_or_equal":
                    occurrence_passed = tool_occurrences >= expected_count
                elif occurrence_comparison_type == "less_than":
                    occurrence_passed = tool_occurrences < expected_count
                elif occurrence_comparison_type == "less_or_equal":
                    occurrence_passed = tool_occurrences <= expected_count
                elif occurrence_comparison_type == "not_equals":
                    occurrence_passed = tool_occurrences != expected_count
                else:
                    return {
                        "passed": False,
                        "error": f"Unsupported occurrence_comparison_type: {occurrence_comparison_type}",
                        "tools_called": tools_called,
                        "missing_tools": missing_tools,
                    }

                passed = passed and occurrence_passed

        return {
            "passed": passed,
            "expected_tools": expected_tools,
            "tools_called": tools_called,
            "missing_tools": missing_tools,
            "minimum_tool_calls": minimum_tool_calls,
            "actual_tool_calls": len(tools_called),
            "occurrence_check": occurrence_details,
            "call_bounds": call_bounds,
            "exclude_errored_calls": exclude_errored_calls,
            "errored_calls": errored_calls,
        }

    async def _execute_sql_query(
        self,
        query: str,
        gym_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Execute SQL query on the CORRECT gym's database"""
        try:
            logger.info(f"Executing SQL query: {query}")

            protocol_client = self._get_protocol_client_for_gym(gym_name)

            if self.execution_mode == "openenv":

                state_result = await protocol_client.get_state(verify_queries=[query])

                if not state_result.get("success"):
                    return {
                        "success": False,
                        "error": state_result.get("error", "State query failed"),
                    }

                state_data = state_result.get("data", {})
                verification_results = state_data.get("verification_results", [])

                if not verification_results:
                    return {
                        "success": False,
                        "error": "No verification results returned from /state",
                    }

                query_result = verification_results[0]

                if "error" in query_result:
                    return {
                        "success": False,
                        "error": query_result.get("error", "Query execution failed"),
                    }

                return {"success": True, "result": query_result.get("result", [])}

            else:

                headers = {
                    "Content-Type": "application/json",
                    "x-database-id": protocol_client.database_id,
                }

                if protocol_client.context and isinstance(
                    protocol_client.context, dict
                ):
                    for key, value in protocol_client.context.items():
                        header_key = f"x-{key.lower().replace('_', '-')}"
                        headers[header_key] = str(value)

                payload = {"query": query}

                async with httpx.AsyncClient(timeout=30.0) as client:
                    response = await client.post(
                        f"{protocol_client.base_url}/api/sql-runner",
                        json=payload,
                        headers=headers,
                    )

                    if response.status_code == 200:
                        result_data = response.json()
                        return {"success": True, "result": result_data}

                    return {
                        "success": False,
                        "error": f"SQL runner failed: HTTP {response.status_code}",
                    }

        except Exception as e:
            logger.error(f"SQL query execution failed: {e}")
            return {"success": False, "error": str(e)}

    def _extract_value_from_sql_result(self, result: dict) -> Any:
        """Extract the actual value from SQL query result (matches production implementation)"""
        if not result:
            return None

        result_data = result.get("result", {})

        if isinstance(result_data, list):
            if result_data:

                if (
                    len(result_data) == 1
                    and isinstance(result_data[0], dict)
                    and len(result_data[0]) == 1
                ):
                    return list(result_data[0].values())[0]

                elif len(result_data) == 1:
                    return result_data[0]

            return result_data

        if isinstance(result_data, dict):

            if "data" in result_data:
                data = result_data["data"]
                if isinstance(data, list) and data:

                    if (
                        len(data) == 1
                        and isinstance(data[0], dict)
                        and len(data[0]) == 1
                    ):
                        return list(data[0].values())[0]

                    elif len(data) == 1:
                        return data[0]

                return data

            elif "rows" in result_data:
                rows = result_data["rows"]
                if isinstance(rows, list) and rows:

                    if (
                        len(rows) == 1
                        and isinstance(rows[0], dict)
                        and len(rows[0]) == 1
                    ):
                        return list(rows[0].values())[0]

                    elif (
                        len(rows) == 1
                        and isinstance(rows[0], list)
                        and len(rows[0]) == 1
                    ):
                        return rows[0][0]

                    elif len(rows) == 1:
                        return rows[0]

                return rows

            elif "content" in result_data:
                content = result_data["content"]
                if isinstance(content, list) and content:
                    for item in content:
                        if isinstance(item, dict) and item.get("type") == "text":
                            return item.get("text", result_data)
                return content

            elif "result" in result_data:
                return result_data["result"]

        return result_data

    def _compare_values(
        self, actual: Any, expected: Any, comparison_type: str
    ) -> Dict[str, Any]:
        """Compare actual vs expected values"""
        try:
            if comparison_type == "equals":
                passed = actual == expected
            elif comparison_type == "not_equals":
                passed = actual != expected
            elif comparison_type == "greater_than":
                passed = actual > expected
            elif comparison_type == "greater_or_equal":
                passed = actual >= expected
            elif comparison_type == "less_than":
                passed = actual < expected
            elif comparison_type == "less_or_equal":
                passed = actual <= expected
            elif comparison_type == "contains":
                passed = expected in str(actual)
            else:
                return {
                    "passed": False,
                    "details": f"Unknown comparison type: {comparison_type}",
                }

            return {
                "passed": passed,
                "details": f"Comparison {comparison_type}: {actual} vs {expected}",
            }

        except Exception as e:
            return {"passed": False, "details": f"Comparison error: {e}"}

    def _extract_llm_content(self, model_response: Dict[str, Any]) -> str:
        """Extract text content from LLM response (string or provider content blocks)."""
        content = model_response.get("content")
        if content is None:
            content = model_response.get("text")
        if content is None:
            content = model_response.get("response")
        if content is None:
            return str(model_response)
        if isinstance(content, list):
            parts: list[str] = []
            for block in content:
                if isinstance(block, dict):
                    text = block.get("text")
                    if text is not None:
                        parts.append(str(text))
                elif isinstance(block, str):
                    parts.append(block)
            return "\n".join(parts)
        return str(content)

    async def _compare_with_llm(
        self,
        sql_result: Dict[str, Any],
        llm_response: str,
        comparison_prompt: str,
        minimum_score: int,
    ) -> Dict[str, Any]:
        """Use LLM as judge to compare SQL result with LLM response"""
        from langchain_core.messages import SystemMessage, HumanMessage

        system_prompt = """You are an AI judge evaluating the quality and accuracy of an AI assistant's response.
Compare the database query result with the AI's response and rate how well they match.
Provide a score from 1-10 where:
- 1-3: Poor match, incorrect or missing information
- 4-6: Partial match, some correct information
- 7-8: Good match, mostly correct
- 9-10: Excellent match, fully accurate

Respond with ONLY a JSON object in this format:
{
  "score": <number 1-10>,
  "reasoning": "<brief explanation>"
}"""

        sql_result_str = json.dumps(sql_result.get("result", {}), indent=2)

        user_prompt = f"""Database Query Result:
{sql_result_str}

AI Assistant Response:
{llm_response}

Comparison Task:
{comparison_prompt}

Please provide your judgment as JSON."""

        messages = [
            SystemMessage(content=system_prompt),
            HumanMessage(content=user_prompt),
        ]

        try:
            response_text = await self._invoke_judge_llm(
                messages, response_usable=_judge_json_parseable
            )
            if not _judge_json_parseable(response_text):
                return {
                    "passed": False,
                    "error": (
                        "Judge comparison failed: empty or unparseable JSON "
                        f"after {max(1, _env_int('TH_JUDGE_JSON_RETRY_MAX', 3))} attempt(s)"
                    ),
                }

            response_text = _extract_fenced_json_text(response_text)

            judge_result = json.loads(response_text)
            score = judge_result.get("score", 0)
            reasoning = judge_result.get("reasoning", "")

            passed = score >= minimum_score

            return {
                "passed": passed,
                "score": score,
                "minimum_score": minimum_score,
                "reasoning": reasoning,
                "sql_result": sql_result_str,
                "llm_response": llm_response,
            }

        except Exception as e:
            logger.error(f"LLM judge comparison failed: {e}")
            return {"passed": False, "error": f"Judge comparison failed: {e}"}


# ── carried over from benchmark_runner.py ──────────────────────────────────
# _normalize_ai_content reaches for _safe_json_preview on reasoning blocks; the
# class extraction left both it and its helpers behind. Defined here (after the
# extracted chunks) because they depend on _env_int.
_TOOL_JSON_PREVIEW_MAX = _env_int("TH_JSON_PREVIEW_MAX", 4000)


def _truncate_text(text: str, max_len: int) -> str:
    if not text:
        return ""
    if len(text) <= max_len:
        return text
    return f"{text[:max_len]}… ({len(text)} chars total)"


def _safe_json_preview(obj: Any, max_len: int = _TOOL_JSON_PREVIEW_MAX) -> str:
    try:
        return _truncate_text(
            json.dumps(obj, default=str, ensure_ascii=False), max_len
        )
    except (TypeError, ValueError):
        return _truncate_text(str(obj), max_len)
